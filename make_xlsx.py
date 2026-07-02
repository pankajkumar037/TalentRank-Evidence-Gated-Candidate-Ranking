#!/usr/bin/env python3
"""
make_xlsx.py — Mirror submission.csv into submission.xlsx.

    python make_xlsx.py --csv submission.csv --out submission.xlsx

The portal asks for BOTH the .csv and the .xlsx, so this produces the .xlsx counterpart
of the ranking. rank.py calls write_xlsx() automatically on every run; this script is
for regenerating the .xlsx standalone from an existing CSV. The xlsx write is kept
separate/lazy so the core ranking path stays reproducible even without openpyxl.

Reads the exact 4 columns the CSV already contains (candidate_id, rank, score,
reasoning), typing rank as int and score as float so Excel sorts them correctly.
"""
from __future__ import annotations

import argparse
import csv

HEADER = ["candidate_id", "rank", "score", "reasoning"]


def write_xlsx(csv_path: str, out_path: str) -> int:
    """Mirror a submission CSV into an XLSX. Returns the number of data rows written.

    Importing openpyxl lazily keeps the core ranking path (rank.py) reproducible even
    on a minimal environment without openpyxl installed.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    with open(csv_path, encoding="utf-8", newline="") as f:
        reader = csv.reader(f)
        header = next(reader)
        if header != HEADER:
            raise ValueError(f"Unexpected CSV header {header!r}; expected {HEADER!r}")
        rows = [row for row in reader if any(cell.strip() for cell in row)]

    wb = Workbook()
    ws = wb.active
    ws.title = "ranking"

    ws.append(HEADER)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for cid, rank, score, reasoning in rows:
        ws.append([cid, int(rank), float(score), reasoning])

    widths = {"candidate_id": 16, "rank": 6, "score": 12, "reasoning": 110}
    for i, name in enumerate(HEADER, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths[name]
    ws.freeze_panes = "A2"

    wb.save(out_path)
    return len(rows)


def main(argv=None):
    ap = argparse.ArgumentParser(description="submission.csv -> submission.xlsx (review only)")
    ap.add_argument("--csv", default="submission.csv", help="input submission CSV")
    ap.add_argument("--out", default="submission.xlsx", help="output XLSX path")
    args = ap.parse_args(argv)

    n = write_xlsx(args.csv, args.out)
    print(f"[done] wrote {n} rows to {args.out} (submit this alongside the .csv)")


if __name__ == "__main__":
    main()
